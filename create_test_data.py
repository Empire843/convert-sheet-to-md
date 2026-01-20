"""
Script để tạo file Excel và CSV mẫu cho testing
"""
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

def create_excel_sample():
    """Tạo file Excel mẫu với nhiều sheet và ảnh"""
    wb = openpyxl.Workbook()

    # Sheet 1: Dữ liệu nhân viên
    ws1 = wb.active
    ws1.title = 'Nhân Viên'
    ws1.append(['Mã NV', 'Họ Tên', 'Phòng Ban', 'Lương'])
    ws1.append(['NV001', 'Nguyễn Văn A', 'IT', 15000000])
    ws1.append(['NV002', 'Trần Thị B', 'Marketing', 12000000])
    ws1.append(['NV003', 'Lê Văn C', 'HR', 10000000])

    # Sheet 2: Sản phẩm
    ws2 = wb.create_sheet('Sản Phẩm')
    ws2.append(['Mã SP', 'Tên Sản Phẩm', 'Giá', 'Số Lượng'])
    ws2.append(['SP001', 'Laptop Dell XPS', 25000000, 10])
    ws2.append(['SP002', 'iPhone 15 Pro', 30000000, 5])
    ws2.append(['SP003', 'Samsung S24', 20000000, 8])

    # Sheet 3: Dự án
    ws3 = wb.create_sheet('Dự Án')
    ws3.append(['Mã DA', 'Tên Dự Án', 'Ngày Bắt Đầu', 'Trạng Thái'])
    ws3.append(['DA001', 'Website Ecommerce', '2024-01-01', 'Đang thực hiện'])
    ws3.append(['DA002', 'Mobile App', '2024-02-01', 'Hoàn thành'])

    # Tạo ảnh mẫu
    try:
        img_pil = PILImage.new('RGB', (200, 100), color='blue')
        img_pil.save('/tmp/test_image.png')
        img = Image('/tmp/test_image.png')
        img.width = 200
        img.height = 100
        ws1.add_image(img, 'F2')
    except Exception as e:
        print(f"Không thể thêm ảnh: {e}")

    # Lưu file
    wb.save('input/sample_data.xlsx')
    print('✓ Đã tạo file sample_data.xlsx')


def create_csv_sample():
    """Tạo file CSV mẫu"""
    data = {
        'Mã KH': ['KH001', 'KH002', 'KH003', 'KH004'],
        'Tên Khách Hàng': ['Công ty ABC', 'Công ty XYZ', 'Doanh nghiệp 123', 'Tập đoàn DEF'],
        'Email': ['abc@example.com', 'xyz@example.com', '123@example.com', 'def@example.com'],
        'Số Điện Thoại': ['0901234567', '0912345678', '0923456789', '0934567890'],
        'Địa Chỉ': ['Hà Nội', 'Hồ Chí Minh', 'Đà Nẵng', 'Cần Thơ']
    }

    df = pd.DataFrame(data)
    df.to_csv('input/khachhang.csv', index=False, encoding='utf-8')
    print('✓ Đã tạo file khachhang.csv')


if __name__ == '__main__':
    print('Tạo file test mẫu...')
    create_excel_sample()
    create_csv_sample()
    print('\nHoàn thành!')
