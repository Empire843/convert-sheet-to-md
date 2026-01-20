"""
Module xử lý file Excel
Đọc và parse file Excel (.xlsx, .xls) thành DataFrame
"""
import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Class để xử lý file Excel"""
    
    def __init__(self, file_path: str):
        """
        Khởi tạo ExcelProcessor
        
        Args:
            file_path: Đường dẫn đến file Excel
        """
        self.file_path = file_path
        self.file_extension = file_path.lower().split('.')[-1]
    
    def get_sheet_names(self) -> List[str]:
        """
        Lấy danh sách tên tất cả các sheet trong Excel
        
        Returns:
            List tên các sheet
        """
        try:
            if self.file_extension == 'xlsx':
                # Dùng openpyxl cho .xlsx
                workbook = load_workbook(self.file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                return sheet_names
            else:
                # Dùng pandas cho .xls
                excel_file = pd.ExcelFile(self.file_path)
                return excel_file.sheet_names
        except Exception as e:
            logger.error(f"Lỗi khi lấy danh sách sheet: {str(e)}")
            return []
    
    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        """
        Đọc một sheet thành DataFrame
        
        Args:
            sheet_name: Tên sheet cần đọc
            
        Returns:
            DataFrame chứa dữ liệu của sheet
        """
        try:
            if self.file_extension == 'xlsx':
                # Đọc .xlsx với openpyxl engine
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    engine='openpyxl'
                )
            else:
                # Đọc .xls với xlrd engine
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=sheet_name,
                    engine='xlrd'
                )
            
            # Replace NaN với empty string cho dễ xử lý
            df = df.fillna('')
            
            return df
            
        except Exception as e:
            logger.error(f"Lỗi khi đọc sheet '{sheet_name}': {str(e)}")
            return pd.DataFrame()
    
    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Đọc tất cả các sheet thành dict of DataFrames
        
        Returns:
            Dict với key là tên sheet, value là DataFrame
        """
        all_data = {}
        sheet_names = self.get_sheet_names()
        
        for sheet_name in sheet_names:
            logger.info(f"Đang đọc sheet: {sheet_name}")
            df = self.read_sheet(sheet_name)
            all_data[sheet_name] = df
        
        return all_data
    
    def get_sheet_info(self) -> Dict[str, Dict]:
        """
        Lấy thông tin metadata của các sheet
        
        Returns:
            Dict chứa thông tin về số hàng, cột của mỗi sheet
        """
        info = {}
        all_data = self.read_all_sheets()
        
        for sheet_name, df in all_data.items():
            info[sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns)
            }
        
        return info
