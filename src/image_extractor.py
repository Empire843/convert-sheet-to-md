"""
Module trích xuất ảnh từ Excel
Lấy images và charts từ Excel workbook
"""
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from typing import List, Tuple, Dict
import logging

logger = logging.getLogger(__name__)


class ImageExtractor:
    """Class để trích xuất ảnh từ Excel"""
    
    def __init__(self, file_path: str, output_dir: str):
        """
        Khởi tạo ImageExtractor
        
        Args:
            file_path: Đường dẫn đến file Excel
            output_dir: Thư mục lưu ảnh
        """
        self.file_path = file_path
        self.output_dir = output_dir
        
        # Tạo thư mục output nếu chưa tồn tại
        os.makedirs(output_dir, exist_ok=True)
    
    def extract_images(self, sheet_name: str) -> List[Dict[str, str]]:
        """
        Trích xuất tất cả ảnh từ một sheet
        
        Args:
            sheet_name: Tên sheet cần trích xuất ảnh
            
        Returns:
            List các dict chứa thông tin ảnh: {'path': str, 'alt': str}
        """
        try:
            workbook = load_workbook(self.file_path)
            
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' không tồn tại")
                return []
            
            worksheet = workbook[sheet_name]
            images_info = []
            
            # Kiểm tra xem sheet có ảnh không
            if not hasattr(worksheet, '_images') or not worksheet._images:
                return []
            
            # Trích xuất từng ảnh
            for idx, image in enumerate(worksheet._images, 1):
                try:
                    # Tạo tên file cho ảnh
                    image_filename = f"{sheet_name}_image_{idx}.png"
                    image_path = os.path.join(self.output_dir, image_filename)
                    
                    # Lưu ảnh
                    with open(image_path, 'wb') as f:
                        f.write(image._data())
                    
                    images_info.append({
                        'path': f"./{image_filename}",
                        'alt': f"Image {idx} from {sheet_name}"
                    })
                    
                    logger.info(f"Đã trích xuất ảnh: {image_filename}")
                    
                except Exception as e:
                    logger.error(f"Lỗi khi trích xuất ảnh {idx} từ sheet {sheet_name}: {str(e)}")
            
            workbook.close()
            return images_info
            
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất ảnh từ sheet {sheet_name}: {str(e)}")
            return []
    
    def extract_all_images(self) -> Dict[str, List[Dict[str, str]]]:
        """
        Trích xuất tất cả ảnh từ tất cả các sheet
        
        Returns:
            Dict với key là tên sheet, value là list thông tin ảnh
        """
        try:
            workbook = load_workbook(self.file_path)
            all_images = {}
            
            for sheet_name in workbook.sheetnames:
                images = self.extract_images(sheet_name)
                if images:
                    all_images[sheet_name] = images
            
            workbook.close()
            return all_images
            
        except Exception as e:
            logger.error(f"Lỗi khi trích xuất ảnh: {str(e)}")
            return {}
